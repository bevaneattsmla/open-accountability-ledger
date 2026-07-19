#!/usr/bin/env python3
"""Verification gate for the Open Accountability Ledger repo. Run from anywhere.
Must end 'RESULT: all checks passed' before anything ships."""
import csv, json, pathlib, sys, re
import jsonschema

ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA, VIEWS = ROOT/"data", ROOT/"views"
reg = json.loads((DATA/"open_accountability_ledger.json").read_text(encoding="utf-8"))
schema = json.loads((DATA/"open_accountability_ledger.schema.json").read_text(encoding="utf-8"))
manifest = json.loads((DATA/"manifest.json").read_text(encoding="utf-8"))
entries = reg["entries"]
fails = []
def check(cond, msg):
    print(("  ok  " if cond else " FAIL ") + msg)
    if not cond: fails.append(msg)

print("1. JSON validates against schema")
try:
    jsonschema.validate(reg, schema); check(True, "register conforms to schema")
except jsonschema.ValidationError as e:
    check(False, f"schema error: {e.message} at {list(e.absolute_path)}")

print("2. HOLD rule correctly derived on every entry")
for e in entries:
    want = not (e["status"] in ("UNVERIFIED","LIVE") or e["confidence"]=="LOW")
    check(e["public_use"]==want, f"{e['id']} public_use={e['public_use']} (rule wants {want})")
    check(bool(e["hold_reason"]) == (not want), f"{e['id']} hold_reason presence matches")

print("3. ids unique and well-formed")
ids=[e["id"] for e in entries]
check(len(ids)==len(set(ids)), "all ids unique")
for e in entries:
    check(e["id"][0]==e["lens"], f"{e['id']} id prefix matches lens")

print("4. Headline numbers reproduce the demonstrator")
total=len(entries)
done=sum(1 for e in entries if e["status"]=="DONE")
hold=sum(1 for e in entries if not e["public_use"])
by={k:sum(1 for e in entries if e['lens']==k) for k in 'ABCD'}
check(total==29, f"29 entries (got {total})")
check(done==0, f"fully acted on = 0 of 29 (got {done})")
heldids=[e['id'] for e in entries if not e['public_use']]
HELD={'A6','A7','A8','A9','A10','A11','A12'}
check(hold==7 and set(heldids)==HELD, f"7 on HOLD ({sorted(HELD)}) (got {hold}: {heldids})")
check(by=={'A':12,'B':4,'C':4,'D':9}, f"lens split A12/B4/C4/D9 (got {by})")

print("5. House rule: no banned punctuation anywhere in the register")
BLOB=json.dumps(reg, ensure_ascii=False)
for ch,label in [("—","em-dash"),("→","arrow"),("·","middot"),("–","en-dash")]:
    check(ch not in BLOB, f"no {label} in register")

print("6. Manifest is a valid cache-buster")
check(isinstance(manifest.get("version"), int) and manifest["version"]>=1, f"manifest.version is a positive int ({manifest.get('version')})")
check("open_accountability_ledger.json" in manifest.get("files",[]), "manifest lists the register file")

print("7. CSV mirror round-trips the JSON")
with (VIEWS/"open_accountability_ledger.csv").open(encoding="utf-8-sig") as f:
    rows=list(csv.DictReader(f))
check(len(rows)==total, f"CSV row count matches ({len(rows)})")
mismatch=sum(1 for e,row in zip(entries,rows) for k in ("id","title","status","confidence","source","amount","year") if str(e.get(k,""))!=row.get(k,""))
check(mismatch==0, f"CSV text fields match JSON ({mismatch} mismatches)")

print("8. xlsx opens with all entries + sheets")
try:
    from openpyxl import load_workbook
    wb=load_workbook(VIEWS/"open_accountability_ledger.xlsx")
    check(wb["Entries"].max_row-3==total, f"xlsx Entries has {total} data rows")
    check({"Entries","Self-commitments","Lenses & legend"}<=set(wb.sheetnames), "all three sheets present")
except Exception as ex:
    check(False, f"xlsx error: {ex}")

print()
if fails: print(f"RESULT: {len(fails)} FAILURE(S)"); sys.exit(1)
print("RESULT: all checks passed")
