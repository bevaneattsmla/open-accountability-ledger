#!/usr/bin/env python3
"""
generate_views.py — derive human-editable views from the canonical JSON register.

Reads open_accountability_ledger.json (the single source of truth) and writes:
  - open_accountability_ledger.csv   : flat, one row per entry, opens in any spreadsheet
  - open_accountability_ledger.xlsx  : styled editing workbook with status/confidence dropdowns

JSON is canonical. These are generated views — regenerate after editing the JSON.
"""
import csv, json, pathlib
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
HERE = ROOT / "views"          # views are written here
DATA = ROOT / "data"           # register is read from here
HERE.mkdir(exist_ok=True)
reg = json.loads((DATA / "open_accountability_ledger.json").read_text(encoding="utf-8"))
entries = reg["entries"]

COLS = ["id","lens","title","what","outcome","status","amount","year","source","confidence","public_use","hold_reason","notes"]
DERIVED = {"public_use","hold_reason"}  # auto-computed by build_register — do not hand-edit

# ---- CSV -------------------------------------------------------------------
csv_path = HERE / "open_accountability_ledger.csv"
with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
    w = csv.DictWriter(f, fieldnames=COLS)
    w.writeheader()
    for e in entries:
        w.writerow({c: e.get(c, "") for c in COLS})

# ---- xlsx ------------------------------------------------------------------
NAVY="1B2A4A"; GREEN="0A5D2E"; AMBER="C98500"; RED="C0392B"; GREY="8A938D"
HOLD_FILL = PatternFill("solid", fgColor="FBEFE9")   # soft red tint for HELD rows
HDR_FILL  = PatternFill("solid", fgColor=NAVY)
DER_FILL  = PatternFill("solid", fgColor="EEF0F2")   # grey for derived columns
thin = Side(style="thin", color="D8DAD5")
BORDER = Border(left=thin,right=thin,top=thin,bottom=thin)

wb = Workbook()

# --- Sheet 1: Entries ---
ws = wb.active
ws.title = "Entries"
# title band
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(COLS))
c = ws.cell(1,1, f"Open Accountability Ledger — register v{reg['version']} · {reg['status']}")
c.font = Font(bold=True, color="FFFFFF", size=12); c.fill = HDR_FILL
c.alignment = Alignment(vertical="center", indent=1)
ws.row_dimensions[1].height = 24
note = ws.cell(2,1, "JSON is the source of truth. Columns public_use and hold_reason are AUTO-derived (status UNVERIFIED or confidence LOW → held). Edit status/confidence, not the derived columns; then regenerate.")
note.font = Font(italic=True, color=GREY, size=9)
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=len(COLS))

HROW = 3
for j,col in enumerate(COLS, start=1):
    cell = ws.cell(HROW, j, col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = DER_FILL if col in DERIVED else HDR_FILL
    if col in DERIVED: cell.font = Font(bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER

for i,e in enumerate(entries):
    r = HROW + 1 + i
    held = not e["public_use"]
    for j,col in enumerate(COLS, start=1):
        v = e.get(col, "")
        if col == "public_use": v = "yes" if v else "HOLD"
        cell = ws.cell(r, j, v)
        cell.alignment = Alignment(vertical="top", wrap_text=col in ("title","what","outcome","source","hold_reason","notes"))
        cell.border = BORDER
        cell.font = Font(size=10)
        if held: cell.fill = HOLD_FILL
        elif col in DERIVED: cell.fill = DER_FILL
        if col=="confidence":
            cell.font = Font(size=10, bold=True,
                color={"HIGH":GREEN,"MEDIUM":AMBER,"LOW":RED}.get(v, "000000"))
        if col=="public_use":
            cell.font = Font(size=10, bold=True, color=RED if held else GREEN)

# widths
W = {"id":6,"lens":6,"title":30,"what":42,"outcome":42,"status":12,"amount":18,
     "year":14,"source":34,"confidence":12,"public_use":11,"hold_reason":34,"notes":34}
for j,col in enumerate(COLS, start=1):
    ws.column_dimensions[get_column_letter(j)].width = W[col]
ws.freeze_panes = f"A{HROW+1}"
ws.auto_filter.ref = f"A{HROW}:{get_column_letter(len(COLS))}{HROW+len(entries)}"

# dropdowns on status + confidence
last = HROW + len(entries)
dv_status = DataValidation(type="list",
    formula1='"DONE,PARTIAL,IGNORED,RECURRING,OVERDUE,UNVERIFIED"', allow_blank=False)
dv_conf = DataValidation(type="list",
    formula1='"HIGH,MEDIUM,LOW"', allow_blank=False)
ws.add_data_validation(dv_status); ws.add_data_validation(dv_conf)
sc = get_column_letter(COLS.index("status")+1)
cc = get_column_letter(COLS.index("confidence")+1)
dv_status.add(f"{sc}{HROW+1}:{sc}{last}")
dv_conf.add(f"{cc}{HROW+1}:{cc}{last}")

# --- Sheet 2: Self-commitments ---
ws2 = wb.create_sheet("Self-commitments")
ws2.cell(1,1,"First, the reform holds itself to its own standard.").font = Font(bold=True, color=GREEN, size=12)
for j,h in enumerate(["id","commitment","detail","status"], start=1):
    hc = ws2.cell(3,j,h); hc.font = Font(bold=True, color="FFFFFF"); hc.fill = HDR_FILL
for i,s in enumerate(reg.get("self_commitments", [])):
    r=4+i
    for j,k in enumerate(["id","text","detail","status"], start=1):
        ws2.cell(r,j,s.get(k,"")).alignment = Alignment(vertical="top", wrap_text=True)
for j,w in enumerate([6,58,34,14], start=1):
    ws2.column_dimensions[get_column_letter(j)].width = w

# --- Sheet 3: Legend ---
ws3 = wb.create_sheet("Lenses & legend")
def block(ws, row, title, mapping):
    tc = ws.cell(row,1,title); tc.font = Font(bold=True, color=NAVY, size=11); row+=1
    for k,v in mapping.items():
        ws.cell(row,1,k).font = Font(bold=True)
        ws.cell(row,2,v).alignment = Alignment(wrap_text=True, vertical="top")
        row+=1
    return row+1
r=1
lens_map = {f"{k} · {reg['lenses'][k]['name']}": reg['lenses'][k]['blurb'] for k in reg['lens_order']}
r=block(ws3,r,"Lenses", lens_map)
r=block(ws3,r,"Status values", reg.get("status_values",{}))
r=block(ws3,r,"Confidence values", reg.get("confidence_values",{}))
r=block(ws3,r,"The discipline", reg.get("discipline",{}))
ws3.column_dimensions["A"].width = 26
ws3.column_dimensions["B"].width = 90

xlsx_path = HERE / "open_accountability_ledger.xlsx"
wb.save(xlsx_path)
print(f"wrote {csv_path.name} and {xlsx_path.name} from {len(entries)} entries")
